#!/usr/bin/env python3
"""
Import Excel survey data into PocketBase.
Reads backup.xlsx and imports farmers, household_members, land_plots,
income, sustainability_certs, and demo_models.
"""

import sys
import os
import time
import json
import requests
import openpyxl

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

# ── Configuration ──────────────────────────────────────────────────────────────
PB_URL = "https://tcn.lvtcenter.it.com"
PB_EMAIL = "admin@lvtcenter.it.com"
PB_PASSWORD = "Admin12345#"
EXCEL_PATH = r"C:\Users\User\AppData\Local\Temp\backup.xlsx"

# ── Helpers ────────────────────────────────────────────────────────────────────

def safe_float(val, default=None):
    """Convert value to float safely."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip().replace(",", "")
        if s == "" or s.lower() in ("không", "none", "n/a", "-"):
            return default
        return float(s)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=None):
    """Convert value to int safely."""
    f = safe_float(val)
    if f is None:
        return default
    return int(f)


def vn_bool(val):
    """Map Vietnamese Có → True, anything else → False."""
    if val is None:
        return False
    return str(val).strip().lower() in ("có", "co", "rồi", "roi")


def clean_str(val, default=""):
    """Convert to stripped string."""
    if val is None:
        return default
    return str(val).strip()


def read_sheet_rows(wb, sheet_name):
    """Read all data rows (skip header) as list of dicts {col_num: value}."""
    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=False)):
        if i == 0:
            continue  # skip header
        record = {}
        for cell in row:
            if cell is not None and hasattr(cell, 'column'):
                record[cell.column] = cell.value
        rows.append(record)
    return rows


# ── PocketBase API helpers ─────────────────────────────────────────────────────

class PBClient:
    def __init__(self, base_url, email, password):
        self.base_url = base_url.rstrip("/")
        self.token = None
        self.authenticate(email, password)

    def authenticate(self, email, password):
        url = f"{self.base_url}/api/collections/_superusers/auth-with-password"
        r = requests.post(url, json={"identity": email, "password": password})
        r.raise_for_status()
        self.token = r.json()["token"]
        print(f"Authenticated as superuser.")

    def headers(self):
        return {"Authorization": self.token}

    def create_record(self, collection, data):
        url = f"{self.base_url}/api/collections/{collection}/records"
        r = requests.post(url, json=data, headers=self.headers())
        if r.status_code not in (200, 201):
            raise Exception(f"Create failed [{r.status_code}]: {r.text[:500]}")
        return r.json()

    def list_records(self, collection, page=1, per_page=500, filter_str=""):
        url = f"{self.base_url}/api/collections/{collection}/records"
        params = {"page": page, "perPage": per_page}
        if filter_str:
            params["filter"] = filter_str
        r = requests.get(url, params=params, headers=self.headers())
        r.raise_for_status()
        return r.json()

    def get_all_records(self, collection, filter_str=""):
        """Fetch all records across pages."""
        all_items = []
        page = 1
        while True:
            data = self.list_records(collection, page=page, per_page=500, filter_str=filter_str)
            all_items.extend(data.get("items", []))
            if page >= data.get("totalPages", 1):
                break
            page += 1
        return all_items

    def delete_record(self, collection, record_id):
        url = f"{self.base_url}/api/collections/{collection}/records/{record_id}"
        r = requests.delete(url, headers=self.headers())
        if r.status_code not in (200, 204):
            raise Exception(f"Delete failed [{r.status_code}]: {r.text[:300]}")

    def collection_exists(self, name):
        url = f"{self.base_url}/api/collections/{name}"
        r = requests.get(url, headers=self.headers())
        return r.status_code == 200

    def create_collection(self, name, fields, list_rule="", view_rule="",
                          create_rule="", update_rule="", delete_rule=""):
        url = f"{self.base_url}/api/collections"
        payload = {
            "name": name,
            "type": "base",
            "listRule": list_rule,
            "viewRule": view_rule,
            "createRule": create_rule,
            "updateRule": update_rule,
            "deleteRule": delete_rule,
            "fields": fields,
        }
        r = requests.post(url, json=payload, headers=self.headers())
        if r.status_code not in (200, 201):
            raise Exception(f"Create collection failed [{r.status_code}]: {r.text[:500]}")
        print(f"  Created collection '{name}'.")
        return r.json()


# ── Ensure 'income_records' collection exists (created by create_collections.py) ──

def ensure_income_collection(pb):
    if pb.collection_exists("income_records"):
        print("  Collection 'income_records' already exists.")
        return
    print("  WARNING: 'income_records' collection not found! Run create_collections.py first.")


# ── Step 1: Import Farmers ─────────────────────────────────────────────────────

def import_farmers(pb, wb):
    print("\n=== Step 1: Importing FARMERS from READABLE_SURVEY_MAIN ===")
    rows = read_sheet_rows(wb, "READABLE_SURVEY_MAIN")
    total = len(rows)
    print(f"  Found {total} rows.")

    # Build mapping of existing farmer_codes to avoid duplicates
    existing = pb.get_all_records("farmers")
    existing_codes = {r["farmer_code"]: r["id"] for r in existing}
    print(f"  {len(existing_codes)} farmers already in PB.")

    farmer_map = dict(existing_codes)  # farmer_code -> pb_id
    imported = 0
    skipped = 0
    errors = 0

    for idx, r in enumerate(rows):
        farmer_code = clean_str(r.get(1))
        if not farmer_code:
            skipped += 1
            continue

        if farmer_code in farmer_map:
            # Already imported
            skipped += 1
            if (idx + 1) % 50 == 0:
                print(f"  Importing farmers: {idx+1}/{total}...")
            continue

        # Map gender: Excel has "Nam"/"Nữ", PB select has "Nam"/"Nu"/"Khac"
        gender_raw = clean_str(r.get(13))
        if gender_raw == "Nam":
            gender = "Nam"
        elif gender_raw in ("Nữ", "Nu"):
            gender = "Nu"
        else:
            gender = "Khac"

        # Date of birth from year
        year = safe_int(r.get(14))
        dob = f"{year}-01-01" if year else None

        # Economic class mapping
        econ_raw = clean_str(r.get(12))
        econ_map = {
            "Nghèo": "Nghèo",
            "Cận nghèo": "Cận nghèo",
            "Bình thường": "Bình thường",
            "Khá": "Khá",
        }
        economic_class = econ_map.get(econ_raw, None)

        # Phone - convert to string, truncate to 20 chars (PB field max)
        phone_raw = r.get(10)
        phone = ""
        if phone_raw is not None:
            phone = str(int(phone_raw)) if isinstance(phone_raw, (int, float)) else str(phone_raw).strip()
            phone = phone[:20]

        data = {
            "farmer_code": farmer_code,
            "full_name": clean_str(r.get(11)) or "N/A",
            "gender": gender,
            "date_of_birth": dob,
            "phone": phone,
            "village": clean_str(r.get(8)),
            "commune": clean_str(r.get(7)),
            "ethnicity": clean_str(r.get(15)),
            "economic_class": economic_class,
            "cooperative_member": vn_bool(r.get(18)),
            "coffee_years": safe_float(r.get(19)),
            "education": clean_str(r.get(17)),
            "total_farms": safe_int(r.get(45)),
            "province": "Gia Lai",
            "household_head": True,
            "status": "active",
        }

        # Remove None values to avoid PB validation errors
        data = {k: v for k, v in data.items() if v is not None}

        try:
            result = pb.create_record("farmers", data)
            farmer_map[farmer_code] = result["id"]
            imported += 1
        except Exception as e:
            errors += 1
            print(f"  ERROR row {idx+2} ({farmer_code}): {e}")

        if (idx + 1) % 50 == 0:
            print(f"  Importing farmers: {idx+1}/{total}...")

    print(f"  Farmers done: imported={imported}, skipped={skipped}, errors={errors}")
    return farmer_map


# ── Step 2: Import Household Members ───────────────────────────────────────────

def import_household_members(pb, wb, farmer_map):
    print("\n=== Step 2: Importing HOUSEHOLD_MEMBERS from READABLE_HH_members ===")
    rows = read_sheet_rows(wb, "READABLE_HH_members")
    total = len(rows)
    print(f"  Found {total} rows.")

    imported = 0
    skipped = 0
    errors = 0

    for idx, r in enumerate(rows):
        farmer_code = clean_str(r.get(4))
        if not farmer_code:
            skipped += 1
            continue

        farmer_id = farmer_map.get(farmer_code)
        if not farmer_id:
            skipped += 1
            if (idx + 1) % 100 == 0:
                print(f"  Importing household_members: {idx+1}/{total}... (skipped={skipped})")
            continue

        data = {
            "farmer_id": farmer_id,
            "member_name": clean_str(r.get(8)) or "N/A",
            "gender": clean_str(r.get(5)),
            "birth_year": safe_int(r.get(6)),
            "relation_to_head": clean_str(r.get(9)),
            "education": clean_str(r.get(10)),
            "coffee_participation": vn_bool(r.get(11)),
            "production_stages": clean_str(r.get(12)),
        }
        data = {k: v for k, v in data.items() if v is not None}

        try:
            pb.create_record("household_members", data)
            imported += 1
        except Exception as e:
            errors += 1
            print(f"  ERROR row {idx+2} ({farmer_code}): {e}")

        if (idx + 1) % 100 == 0:
            print(f"  Importing household_members: {idx+1}/{total}...")

    print(f"  Household members done: imported={imported}, skipped={skipped}, errors={errors}")
    return imported


# ── Step 3: Import Land Plots ──────────────────────────────────────────────────

def import_land_plots(pb, wb, farmer_map):
    print("\n=== Step 3: Importing LAND_PLOTS from READABLE_LAND_PLOTS ===")
    rows = read_sheet_rows(wb, "READABLE_LAND_PLOTS")
    total = len(rows)
    print(f"  Found {total} rows.")

    imported = 0
    skipped = 0
    errors = 0

    for idx, r in enumerate(rows):
        farmer_code = clean_str(r.get(4))
        if not farmer_code:
            skipped += 1
            continue

        farmer_id = farmer_map.get(farmer_code)
        if not farmer_id:
            skipped += 1
            if (idx + 1) % 100 == 0:
                print(f"  Importing land_plots: {idx+1}/{total}... (skipped={skipped})")
            continue

        data = {
            "farmer_id": farmer_id,
            "plot_name": clean_str(r.get(6)) or "N/A",
            "area_ha": safe_float(r.get(7)),
            "forest_area": safe_float(r.get(8)),
            "tree_count": safe_int(r.get(9)),
            "intercrop": vn_bool(r.get(10)),
            "intercrop_species": clean_str(r.get(11)),
            "harvest_area": safe_float(r.get(13)),
            "yield_current": safe_float(r.get(14)),
            "ownership_type": clean_str(r.get(15)),
            "start_year": safe_int(r.get(17)),
            "latest_planting": safe_int(r.get(18)),
            "previous_crop": clean_str(r.get(19)),
            "forest_distance_km": safe_float(r.get(21)),
            "gps_mapped": vn_bool(r.get(22)),
        }
        data = {k: v for k, v in data.items() if v is not None}

        try:
            pb.create_record("land_plots", data)
            imported += 1
        except Exception as e:
            errors += 1
            print(f"  ERROR row {idx+2} ({farmer_code}): {e}")

        if (idx + 1) % 100 == 0:
            print(f"  Importing land_plots: {idx+1}/{total}...")

    print(f"  Land plots done: imported={imported}, skipped={skipped}, errors={errors}")
    return imported


# ── Step 4: Import Income ──────────────────────────────────────────────────────

def import_income(pb, wb, farmer_map):
    print("\n=== Step 4: Importing INCOME from READABLE_SURVEY_MAIN ===")
    rows = read_sheet_rows(wb, "READABLE_SURVEY_MAIN")
    total = len(rows)
    print(f"  Found {total} rows.")

    imported = 0
    skipped = 0
    errors = 0

    for idx, r in enumerate(rows):
        farmer_code = clean_str(r.get(1))
        if not farmer_code:
            skipped += 1
            continue

        farmer_id = farmer_map.get(farmer_code)
        if not farmer_id:
            skipped += 1
            continue

        data = {
            "farmer_id": farmer_id,
            "year": 2025,
            "total_income": safe_float(r.get(25)),
            "agri_income_ratio": safe_float(r.get(26)),
            "coffee_revenue": safe_float(r.get(30)),
            "coffee_cost": safe_float(r.get(31)),
            "coffee_net": safe_float(r.get(29)),
            "rice_income": safe_float(r.get(32)),
            "fruit_income": safe_float(r.get(33)),
            "livestock_income": safe_float(r.get(37)),
            "salary_income": safe_float(r.get(38)),
            "business_income": safe_float(r.get(40)),
            "other_income": safe_float(r.get(41)),
            "production_tons": safe_float(r.get(83)),
            "production_value": safe_float(r.get(84)),
        }
        data = {k: v for k, v in data.items() if v is not None}

        try:
            pb.create_record("income_records", data)
            imported += 1
        except Exception as e:
            errors += 1
            print(f"  ERROR row {idx+2} ({farmer_code}): {e}")

        if (idx + 1) % 50 == 0:
            print(f"  Importing income: {idx+1}/{total}...")

    print(f"  Income done: imported={imported}, skipped={skipped}, errors={errors}")
    return imported


# ── Step 5: Import Sustainability Certs ────────────────────────────────────────

def import_sustainability_certs(pb, wb, farmer_map):
    print("\n=== Step 5: Importing SUSTAINABILITY_CERTS from READABLE_SUBTAINABLE_AGRI ===")
    rows = read_sheet_rows(wb, "READABLE_SUBTAINABLE_AGRI")
    total = len(rows)
    print(f"  Found {total} rows.")

    imported = 0
    skipped = 0
    errors = 0

    for idx, r in enumerate(rows):
        farmer_code = clean_str(r.get(5))
        if not farmer_code:
            skipped += 1
            continue

        farmer_id = farmer_map.get(farmer_code)
        if not farmer_id:
            skipped += 1
            continue

        # cert_type: use C3 (cert category) as primary, C6 (detail) as fallback
        cert_type = clean_str(r.get(3))  # e.g. "4C", "Trồng xen canh", "Hữu cơ, Organic"
        cert_detail = clean_str(r.get(6))  # only filled for "khác" (other)
        if not cert_type and cert_detail:
            cert_type = cert_detail
        elif cert_type and cert_detail:
            cert_type = f"{cert_type}: {cert_detail}"

        if not cert_type:
            skipped += 1
            continue

        data = {
            "farmer_id": farmer_id,
            "cert_type": cert_type,
            "family_knows": vn_bool(r.get(7)),
            "who_knows_most": clean_str(r.get(8)),
            "is_beneficial": vn_bool(r.get(9)),
            "practiced_on_land": vn_bool(r.get(10)),
            "is_certified": vn_bool(r.get(11)),
            "who_practices": clean_str(r.get(12)),
        }
        data = {k: v for k, v in data.items() if v is not None}

        try:
            pb.create_record("sustainability_certs", data)
            imported += 1
        except Exception as e:
            errors += 1
            print(f"  ERROR row {idx+2} ({farmer_code}): {e}")

        if (idx + 1) % 50 == 0:
            print(f"  Importing sustainability_certs: {idx+1}/{total}...")

    print(f"  Sustainability certs done: imported={imported}, skipped={skipped}, errors={errors}")
    return imported


# ── Step 6: Create Demo Models ─────────────────────────────────────────────────

def create_demo_models(pb):
    print("\n=== Step 6: Creating DEMO_MODELS ===")

    models = [
        {
            "model_code": "GL01-XP",
            "model_name": "Mô hình GL01 - Đăk Đoa",
            "province": "Gia Lai",
            "status": "active",
            "data_status": "full",
        },
        {
            "model_code": "GL02-XP",
            "model_name": "Mô hình GL02 - Đăk Đoa",
            "province": "Gia Lai",
            "status": "active",
            "data_status": "partial",
        },
        {
            "model_code": "GL03-XP",
            "model_name": "Mô hình GL03 - Placeholder",
            "province": "Gia Lai",
            "status": "planning",
            "data_status": "partial",
        },
        {
            "model_code": "GL04-XP",
            "model_name": "Mô hình GL04 - Đăk Đoa/Kon Gang",
            "province": "Gia Lai",
            "status": "active",
            "data_status": "partial",
        },
        {
            "model_code": "GL05-XP",
            "model_name": "Mô hình GL05 - Đăk Đoa/Kon Gang",
            "province": "Gia Lai",
            "status": "active",
            "data_status": "full",
        },
        {
            "model_code": "GL06-XP",
            "model_name": "Mô hình GL06 - Đăk Đoa",
            "province": "Gia Lai",
            "status": "active",
            "data_status": "partial",
        },
        {
            "model_code": "GL07-XP",
            "model_name": "Mô hình GL07 - Đăk Đoa/Kon Gang",
            "province": "Gia Lai",
            "status": "active",
            "data_status": "full",
        },
        {
            "model_code": "GL08-XP",
            "model_name": "Mô hình GL08 - Placeholder",
            "province": "Gia Lai",
            "status": "planning",
            "data_status": "pending",
        },
        {
            "model_code": "GL09-XP",
            "model_name": "Mô hình GL09 - Placeholder",
            "province": "Gia Lai",
            "status": "planning",
            "data_status": "pending",
        },
    ]

    # Check existing
    existing = pb.get_all_records("demo_models")
    existing_codes = {r["model_code"] for r in existing}

    imported = 0
    skipped = 0
    errors = 0

    for m in models:
        if m["model_code"] in existing_codes:
            print(f"  Skipping {m['model_code']} (already exists).")
            skipped += 1
            continue
        try:
            pb.create_record("demo_models", m)
            imported += 1
            print(f"  Created {m['model_code']}: {m['model_name']}")
        except Exception as e:
            errors += 1
            print(f"  ERROR creating {m['model_code']}: {e}")

    print(f"  Demo models done: imported={imported}, skipped={skipped}, errors={errors}")
    return imported


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("  TCN Excel Data Import to PocketBase")
    print("=" * 70)

    # Check Excel file exists
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    print(f"\nExcel file: {EXCEL_PATH}")
    print(f"PocketBase: {PB_URL}")
    print()

    # Load workbook
    print("Loading Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # Connect to PocketBase
    print("\nConnecting to PocketBase...")
    pb = PBClient(PB_URL, PB_EMAIL, PB_PASSWORD)

    # Ensure income collection exists
    print("\nEnsuring 'income' collection exists...")
    ensure_income_collection(pb)

    # Track summary
    summary = {}
    start_time = time.time()

    # Step 1: Farmers
    farmer_map = import_farmers(pb, wb)
    summary["farmers"] = len(farmer_map)

    # Step 2: Household Members
    summary["household_members"] = import_household_members(pb, wb, farmer_map)

    # Step 3: Land Plots
    summary["land_plots"] = import_land_plots(pb, wb, farmer_map)

    # Step 4: Income
    summary["income_records"] = import_income(pb, wb, farmer_map)

    # Step 5: Sustainability Certs
    summary["sustainability_certs"] = import_sustainability_certs(pb, wb, farmer_map)

    # Step 6: Demo Models
    summary["demo_models"] = create_demo_models(pb)

    # Close workbook
    wb.close()

    # Print summary
    elapsed = time.time() - start_time
    print("\n" + "=" * 70)
    print("  IMPORT SUMMARY")
    print("=" * 70)
    for collection, count in summary.items():
        print(f"  {collection:.<30} {count}")
    print(f"\n  Total time: {elapsed:.1f} seconds")
    print(f"  Farmer code mapping size: {len(farmer_map)}")
    print("=" * 70)
    print("  DONE!")
    print("=" * 70)


if __name__ == "__main__":
    main()
