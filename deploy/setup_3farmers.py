"""
Setup PocketBase with ONLY 3 specific farmers + link them to demo_models.
Also imports household_members, land_plots, income_records for these 3 farmers.
Deletes all other farmers.
"""
import sys
import os
import json
import time
import requests

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

PB_URL = "https://tcn.lvtcenter.it.com"
PB_EMAIL = "admin@lvtcenter.it.com"
PB_PASS = "Admin12345#"

# The 3 target farmer codes and their model assignments
TARGET_FARMERS = {
    "GL01-XP_01-080120261449": "GL01-XP",
    "GL07-XP_08-090120261007": "GL07-XP",
    "GL05-XP_07-090120261212": "GL05-XP",
}

session = requests.Session()


def auth():
    r = session.post(f"{PB_URL}/api/collections/_superusers/auth-with-password",
                     json={"identity": PB_EMAIL, "password": PB_PASS})
    r.raise_for_status()
    token = r.json()["token"]
    session.headers.update({"Authorization": f"Bearer {token}"})
    print(f"Authenticated OK")
    return token


def get_all(collection, filter_str=""):
    items = []
    page = 1
    while True:
        params = {"page": page, "perPage": 500}
        if filter_str:
            params["filter"] = filter_str
        r = session.get(f"{PB_URL}/api/collections/{collection}/records", params=params)
        r.raise_for_status()
        d = r.json()
        items.extend(d.get("items", []))
        if page >= d.get("totalPages", 1):
            break
        page += 1
    return items


def delete_record(collection, rid):
    r = session.delete(f"{PB_URL}/api/collections/{collection}/records/{rid}")
    return r.status_code in (200, 204)


def update_record(collection, rid, data):
    r = session.patch(f"{PB_URL}/api/collections/{collection}/records/{rid}", json=data)
    return r


def create_record(collection, data):
    r = session.post(f"{PB_URL}/api/collections/{collection}/records", json=data)
    if r.status_code not in (200, 201):
        raise Exception(f"Create {collection} failed [{r.status_code}]: {r.text[:300]}")
    return r.json()


def main():
    print("=" * 60)
    print("  Setup 3 Farmers + Link to Demo Models")
    print("=" * 60)
    auth()

    # Step 1: Get all farmers, keep only 3 target ones
    print("\n[STEP 1] Cleaning farmers - keeping only 3 targets...")
    all_farmers = get_all("farmers")
    print(f"  Total farmers in PB: {len(all_farmers)}")

    keep_ids = {}  # farmer_code -> PB id
    delete_ids = []

    for f in all_farmers:
        code = f.get("farmer_code", "")
        if code in TARGET_FARMERS:
            keep_ids[code] = f["id"]
            print(f"  KEEP: {code} -> {f['full_name']} (id={f['id']})")
        else:
            delete_ids.append(f["id"])

    print(f"  Keeping {len(keep_ids)}, deleting {len(delete_ids)} farmers...")

    # Delete related records first (household_members, land_plots, income_records, sustainability_certs)
    # that reference farmers being deleted
    for collection in ["household_members", "land_plots", "income_records", "sustainability_certs"]:
        print(f"\n  Cleaning {collection}...")
        all_recs = get_all(collection)
        deleted = 0
        kept = 0
        for rec in all_recs:
            fid = rec.get("farmer_id", "")
            if fid and fid not in keep_ids.values():
                delete_record(collection, rec["id"])
                deleted += 1
            else:
                kept += 1
        print(f"    {collection}: deleted={deleted}, kept={kept}")

    # Now delete the farmers
    print(f"\n  Deleting {len(delete_ids)} excess farmers...")
    del_count = 0
    for i, fid in enumerate(delete_ids):
        if delete_record("farmers", fid):
            del_count += 1
        if (i + 1) % 50 == 0:
            print(f"    Deleted {i+1}/{len(delete_ids)}...")
    print(f"  Deleted {del_count} farmers.")

    # Step 2: Link farmers to demo_models
    print("\n[STEP 2] Linking farmers to demo_models...")
    models = get_all("demo_models")
    model_map = {m["model_code"]: m for m in models}

    for farmer_code, model_code in TARGET_FARMERS.items():
        farmer_id = keep_ids.get(farmer_code)
        model = model_map.get(model_code)
        if not farmer_id or not model:
            print(f"  SKIP: farmer={farmer_code} or model={model_code} not found")
            continue

        r = update_record("demo_models", model["id"], {"farmer_id": farmer_id})
        if r.status_code == 200:
            print(f"  LINKED: {model_code} -> {farmer_code} (farmer_id={farmer_id})")
        else:
            print(f"  ERROR linking {model_code}: {r.text[:200]}")

    # Step 3: Import data for 3 farmers from Excel
    print("\n[STEP 3] Importing household_members, land_plots, income for 3 farmers...")

    try:
        import openpyxl
        excel_path = os.path.join(os.environ.get("TEMP", ""), "backup.xlsx")
        if not os.path.exists(excel_path):
            print(f"  Excel not found at {excel_path}, skipping import.")
            return

        wb = openpyxl.load_workbook(excel_path, read_only=True)

        # Read headers for reference
        def get_headers(sheet_name):
            ws = wb[sheet_name]
            headers = {}
            for row in ws.iter_rows(max_row=1, values_only=False):
                for cell in row:
                    if cell.value:
                        headers[cell.column] = str(cell.value)
            return headers

        def get_rows(sheet_name):
            ws = wb[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=False)):
                if i == 0:
                    continue
                record = {}
                for cell in row:
                    if cell is not None and hasattr(cell, 'column'):
                        record[cell.column] = cell.value
                rows.append(record)
            return rows

        def sf(val):
            """Safe float"""
            if val is None: return None
            if isinstance(val, (int, float)): return float(val)
            try:
                s = str(val).strip().replace(",", "")
                if s in ("", "-", "None", "N/A", "không"): return None
                return float(s)
            except: return None

        def si(val):
            f = sf(val)
            return int(f) if f is not None else None

        def vb(val):
            if val is None: return False
            return str(val).strip().lower() in ("có", "co", "rồi", "roi")

        def cs(val, default=""):
            if val is None: return default
            return str(val).strip()

        # Check if data already imported for these farmers
        for fcode, fid in keep_ids.items():
            existing_hh = get_all("household_members", f"farmer_id='{fid}'")
            existing_lp = get_all("land_plots", f"farmer_id='{fid}'")
            existing_ir = get_all("income_records", f"farmer_id='{fid}'")
            if existing_hh or existing_lp or existing_ir:
                print(f"  {fcode}: already has data (hh={len(existing_hh)}, lp={len(existing_lp)}, ir={len(existing_ir)}), skipping import")
                continue

            print(f"\n  Importing data for {fcode} (id={fid})...")

            # Household members
            hh_rows = get_rows("READABLE_HH_members")
            hh_count = 0
            for r in hh_rows:
                if cs(r.get(4)) == fcode:
                    data = {
                        "farmer_id": fid,
                        "member_name": cs(r.get(8)) or "N/A",
                        "gender": cs(r.get(5)),
                        "birth_year": si(r.get(6)),
                        "relation_to_head": cs(r.get(9)),
                        "education": cs(r.get(10)),
                        "coffee_participation": vb(r.get(11)),
                        "production_stages": cs(r.get(12)),
                    }
                    data = {k: v for k, v in data.items() if v is not None}
                    try:
                        create_record("household_members", data)
                        hh_count += 1
                    except Exception as e:
                        print(f"    HH error: {e}")
            print(f"    household_members: {hh_count}")

            # Land plots
            lp_rows = get_rows("READABLE_LAND_PLOTS")
            lp_count = 0
            for r in lp_rows:
                if cs(r.get(4)) == fcode:
                    data = {
                        "farmer_id": fid,
                        "plot_name": cs(r.get(6)) or "N/A",
                        "area_ha": sf(r.get(7)),
                        "forest_area": sf(r.get(8)),
                        "tree_count": si(r.get(9)),
                        "intercrop": vb(r.get(10)),
                        "intercrop_species": cs(r.get(11)),
                        "harvest_area": sf(r.get(13)),
                        "yield_current": sf(r.get(14)),
                        "ownership_type": cs(r.get(15)),
                        "start_year": si(r.get(17)),
                        "latest_planting": si(r.get(18)),
                        "previous_crop": cs(r.get(19)),
                        "forest_distance_km": sf(r.get(21)),
                        "gps_mapped": vb(r.get(22)),
                    }
                    data = {k: v for k, v in data.items() if v is not None}
                    try:
                        create_record("land_plots", data)
                        lp_count += 1
                    except Exception as e:
                        print(f"    LP error: {e}")
            print(f"    land_plots: {lp_count}")

            # Income records (from READABLE_SURVEY_MAIN)
            survey_rows = get_rows("READABLE_SURVEY_MAIN")
            for r in survey_rows:
                if cs(r.get(1)) == fcode:
                    data = {
                        "farmer_id": fid,
                        "year": 2025,
                        "total_income": sf(r.get(25)),
                        "agri_income_ratio": sf(r.get(26)),
                        "coffee_revenue": sf(r.get(30)),
                        "coffee_cost": sf(r.get(31)),
                        "coffee_net": sf(r.get(29)),
                        "rice_income": sf(r.get(32)),
                        "fruit_income": sf(r.get(33)),
                        "livestock_income": sf(r.get(37)),
                        "salary_income": sf(r.get(38)),
                        "business_income": sf(r.get(40)),
                        "other_income": sf(r.get(41)),
                        "production_tons": sf(r.get(83)),
                        "production_value": sf(r.get(84)),
                    }
                    data = {k: v for k, v in data.items() if v is not None}
                    try:
                        create_record("income_records", data)
                        print(f"    income_records: 1")
                    except Exception as e:
                        print(f"    IR error: {e}")
                    break

            # Sustainability certs
            cert_rows = get_rows("READABLE_SUBTAINABLE_AGRI")
            cert_count = 0
            for r in cert_rows:
                if cs(r.get(5)) == fcode:
                    data = {
                        "farmer_id": fid,
                        "cert_type": cs(r.get(6)),
                        "family_knows": vb(r.get(7)),
                        "who_knows_most": cs(r.get(8)),
                        "is_beneficial": vb(r.get(9)),
                        "practiced_on_land": vb(r.get(10)),
                        "is_certified": vb(r.get(11)),
                        "who_practices": cs(r.get(12)),
                    }
                    data = {k: v for k, v in data.items() if v is not None}
                    try:
                        create_record("sustainability_certs", data)
                        cert_count += 1
                    except Exception as e:
                        print(f"    CERT error: {e}")
            print(f"    sustainability_certs: {cert_count}")

        wb.close()

    except ImportError:
        print("  openpyxl not installed, skipping Excel import")
    except Exception as e:
        print(f"  Excel import error: {e}")

    # Final summary
    print("\n" + "=" * 60)
    print("  FINAL STATE")
    print("=" * 60)
    for coll in ["farmers", "demo_models", "household_members", "land_plots", "income_records", "sustainability_certs"]:
        recs = get_all(coll)
        print(f"  {coll}: {len(recs)} records")

    # Show model-farmer links
    print("\n  Demo Model -> Farmer links:")
    models = get_all("demo_models")
    for m in sorted(models, key=lambda x: x.get("model_code", "")):
        fid = m.get("farmer_id", "")
        fname = ""
        if fid:
            for f in get_all("farmers", f"id='{fid}'"):
                fname = f.get("full_name", "")
        print(f"    {m['model_code']}: farmer_id={fid or '---'} ({fname})")

    print("\n  DONE!")


if __name__ == "__main__":
    main()
